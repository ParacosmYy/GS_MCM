"""
扫描 result.xlsx 红字单元格 + 钉死附件 4 规模。
输出 data/result_redcells.json 作为问题 4 答题文件的红字保护基线。
"""
from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
DATA = ROOT / "data"
OUT = DATA / "result_redcells.json"


def scan_red_cells(xlsx_path: Path) -> list[dict]:
    wb = load_workbook(xlsx_path)
    red = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                color = c.font.color
                if c.value is None or color is None:
                    continue
                rgb = getattr(color, "rgb", None)
                if rgb and str(rgb).upper().endswith("FF0000"):
                    red.append(
                        {
                            "sheet": ws.title,
                            "coord": c.coordinate,
                            "row": c.row,
                            "col": c.column,
                            "value": c.value,
                        }
                    )
    return red


def assert_attachment4(path: Path) -> None:
    shoot = pd.read_excel(path, sheet_name="射击目标")
    photo = pd.read_excel(path, sheet_name="拍照目标")
    assert len(shoot) == 18, f"射击目标数异常: 期望 18，实际 {len(shoot)}"
    assert len(photo) == 18, f"拍照目标数异常: 期望 18，实际 {len(photo)}"
    assert shoot.iloc[0, 0] == "S01" and shoot.iloc[-1, 0] == "S18", "射击编号不连续 S01-S18"
    assert photo.iloc[0, 0] == "P01" and photo.iloc[-1, 0] == "P18", "拍照编号不连续 P01-P18"
    print("[OK] 附件4 规模断言通过: 18 射击 + 18 拍照")


def main() -> None:
    redcells = scan_red_cells(DATA / "result.xlsx")
    OUT.write_text(json.dumps(redcells, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[OK] 红字单元格 {len(redcells)} 个 -> {OUT.relative_to(ROOT)}")
    for rc in redcells:
        print(f"  {rc['sheet']}!{rc['coord']}  {rc['value']!r}")
    assert_attachment4(DATA / "附件4.xlsx")


if __name__ == "__main__":
    main()
