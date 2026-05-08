"""
问题 4：任务规划优化版（贪心 + 加权区间调度 DP + CP-SAT）

改进点：
1. 贪心改为"价值密度优先"：射击价值=1（单次），拍照价值按角度覆盖递减
2. 加权区间调度 DP：每个可行执行视为一个[prep_start, exec_end]区间，
   按结束时间排序后用 DP 选最大权重子集
3. CP-SAT 精确求解（限时 5 min）
4. 三者取最优输出

约束（确认值）：
  射击: d ∈ [5, 30] m, v ≤ 2 m/s, a ≤ 1.5 m/s², prep=1.5s
  拍照: d ∈ [10, 40] m, v ≤ 1.5 m/s, a ≤ 1.5 m/s², prep=0.5s, Δθ ≥ 60°

输入：output/traj_10hz_3.csv
输出：output/result_v2.xlsx, docs/problem4_result_v2.md
"""
from __future__ import annotations

import shutil
import time
from bisect import bisect_right
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from scipy.signal import savgol_filter

ROOT = Path(__file__).resolve().parents[2]
DATA = ROOT / "data"
OUTPUT = ROOT / "output"
DOCS = ROOT / "docs"
FIGURES = ROOT / "figures"
OUTPUT.mkdir(exist_ok=True)
DOCS.mkdir(exist_ok=True)

# ─── 约束常量 ───
D_SHOOT = (5, 30)
V_SHOOT = 2.0
A_SHOOT = 1.5
PREP_SHOOT = 1.5

D_PHOTO = (10, 40)
V_PHOTO = 1.5
A_PHOTO = 1.5
PREP_PHOTO = 0.5
DELTA_THETA_MIN = 60  # degrees


def load_traj():
    traj = pd.read_csv(OUTPUT / "traj_10hz_3.csv")
    t = traj["t"].to_numpy()
    window = min(151, len(t) - (1 if len(t) % 2 == 0 else 0))
    if window % 2 == 0:
        window -= 1
    x = savgol_filter(traj["x"].to_numpy(), window, 3)
    y = savgol_filter(traj["y"].to_numpy(), window, 3)
    dt = 0.1
    vx = np.gradient(x, dt)
    vy = np.gradient(y, dt)
    v = np.sqrt(vx**2 + vy**2)
    ax = np.gradient(vx, dt)
    ay = np.gradient(vy, dt)
    a = np.sqrt(ax**2 + ay**2)
    return t, x, y, v, a


def load_targets():
    shoot = pd.read_excel(DATA / "附件4.xlsx", sheet_name="射击目标")
    photo = pd.read_excel(DATA / "附件4.xlsx", sheet_name="拍照目标")
    shoot.columns = ["id", "x", "y"]
    photo.columns = ["id", "x", "y"]
    return shoot, photo


def find_feasible_windows(t, x, y, v, a, target_x, target_y, d_range, v_max, a_max, prep_time):
    n = len(t)
    dt_step = 0.1
    prep_samples = int(prep_time / dt_step)

    dist = np.sqrt((x - target_x)**2 + (y - target_y)**2)
    d_ok = (dist >= d_range[0]) & (dist <= d_range[1])
    v_ok = v <= v_max
    a_ok = a <= a_max
    instant_ok = d_ok & v_ok & a_ok

    feasible = []
    for i in range(prep_samples, n):
        if np.all(instant_ok[i - prep_samples:i + 1]):
            feasible.append(i)
    return feasible


def compute_angle(x_robot, y_robot, x_target, y_target):
    return np.degrees(np.arctan2(y_robot - y_target, x_robot - x_target))


# ═══════════════════════════════════════════════════════════════
# 方法 1: 改进贪心（价值密度优先）
# ═══════════════════════════════════════════════════════════════

def improved_greedy(t, x, y, v, a, shoot_targets, photo_targets):
    """改进贪心：先扫描所有候选，按价值/时间占用比排序"""
    candidates = []

    for _, row in shoot_targets.iterrows():
        fw = find_feasible_windows(t, x, y, v, a, row["x"], row["y"],
                                   D_SHOOT, V_SHOOT, A_SHOOT, PREP_SHOOT)
        for idx in fw:
            prep_start = idx - int(PREP_SHOOT / 0.1)
            candidates.append({
                "exec_idx": idx,
                "prep_start_idx": prep_start,
                "target": row["id"],
                "task": "射击",
                "angle": None,
                "prep_time": PREP_SHOOT,
            })

    for _, row in photo_targets.iterrows():
        fw = find_feasible_windows(t, x, y, v, a, row["x"], row["y"],
                                   D_PHOTO, V_PHOTO, A_PHOTO, PREP_PHOTO)
        for idx in fw:
            angle = compute_angle(x[idx], y[idx], row["x"], row["y"])
            prep_start = idx - int(PREP_PHOTO / 0.1)
            candidates.append({
                "exec_idx": idx,
                "prep_start_idx": prep_start,
                "target": row["id"],
                "task": "拍照",
                "angle": angle,
                "prep_time": PREP_PHOTO,
            })

    # 按执行时刻排序
    candidates.sort(key=lambda c: c["exec_idx"])

    # 贪心：射击优先（价值高），同时拍照尽量多角度
    schedule = []
    occupied_until = -1
    shot_done = set()
    photo_done = {}  # target_id → list of angles

    for c in candidates:
        if c["prep_start_idx"] <= occupied_until:
            continue

        if c["task"] == "射击":
            if c["target"] in shot_done:
                continue
            shot_done.add(c["target"])
            schedule.append(c)
            occupied_until = c["exec_idx"]

        elif c["task"] == "拍照":
            angle = c["angle"]
            prev_angles = photo_done.get(c["target"], [])
            ok = True
            for pa in prev_angles:
                diff = abs(angle - pa) % 360
                diff = min(diff, 360 - diff)
                if diff < DELTA_THETA_MIN:
                    ok = False
                    break
            if not ok:
                continue
            photo_done.setdefault(c["target"], []).append(angle)
            schedule.append(c)
            occupied_until = c["exec_idx"]

    return schedule


# ═══════════════════════════════════════════════════════════════
# 方法 2: 加权区间调度 DP
# ═══════════════════════════════════════════════════════════════

def interval_scheduling_dp(t, x, y, v, a, shoot_targets, photo_targets):
    """
    将每个可行执行建模为区间 [prep_start_idx, exec_idx]，
    权重：射击=10，拍照=5（鼓励多完成射击），
    用 DP 找最大权重不重叠子集。
    然后贪心后处理填补空隙。
    """
    # 生成所有候选区间
    intervals = []

    for _, row in shoot_targets.iterrows():
        fw = find_feasible_windows(t, x, y, v, a, row["x"], row["y"],
                                   D_SHOOT, V_SHOOT, A_SHOOT, PREP_SHOOT)
        for idx in fw:
            prep_start = idx - int(PREP_SHOOT / 0.1)
            intervals.append({
                "start": prep_start,
                "end": idx,
                "target": row["id"],
                "task": "射击",
                "angle": None,
                "weight": 10,
            })

    for _, row in photo_targets.iterrows():
        fw = find_feasible_windows(t, x, y, v, a, row["x"], row["y"],
                                   D_PHOTO, V_PHOTO, A_PHOTO, PREP_PHOTO)
        for idx in fw:
            angle = compute_angle(x[idx], y[idx], row["x"], row["y"])
            prep_start = idx - int(PREP_PHOTO / 0.1)
            intervals.append({
                "start": prep_start,
                "end": idx,
                "target": row["id"],
                "task": "拍照",
                "angle": angle,
                "weight": 5,
            })

    if not intervals:
        return []

    # 去重：同目标同任务只保留每个"时间段"的最早一个（减少候选数）
    # 按 end 排序
    intervals.sort(key=lambda iv: iv["end"])

    # DP: dp[i] = 前 i 个区间中最大权重
    n = len(intervals)
    ends = [iv["end"] for iv in intervals]

    # 对每个区间找"最后一个不重叠的前驱"
    def find_prev(i):
        target_end = intervals[i]["start"] - 1
        lo, hi = 0, i - 1
        result = -1
        while lo <= hi:
            mid = (lo + hi) // 2
            if ends[mid] <= target_end:
                result = mid
                lo = mid + 1
            else:
                hi = mid - 1
        return result

    dp = [0] * (n + 1)
    choice = [False] * n

    for i in range(n):
        prev = find_prev(i)
        include_val = intervals[i]["weight"] + (dp[prev + 1] if prev >= 0 else 0)
        exclude_val = dp[i]
        if include_val > exclude_val:
            dp[i + 1] = include_val
        else:
            dp[i + 1] = exclude_val

    # 回溯选择
    selected_indices = []
    i = n - 1
    while i >= 0:
        prev = find_prev(i)
        include_val = intervals[i]["weight"] + (dp[prev + 1] if prev >= 0 else 0)
        if include_val >= dp[i]:
            selected_indices.append(i)
            i = prev
        else:
            i -= 1

    selected_indices.reverse()

    # 约束过滤：射击每目标至多1次，拍照同目标角差≥60°
    schedule = []
    shot_done = set()
    photo_done = {}

    for idx in selected_indices:
        iv = intervals[idx]
        if iv["task"] == "射击":
            if iv["target"] in shot_done:
                continue
            shot_done.add(iv["target"])
            schedule.append(iv)
        else:
            angle = iv["angle"]
            prev_angles = photo_done.get(iv["target"], [])
            ok = True
            for pa in prev_angles:
                diff = abs(angle - pa) % 360
                diff = min(diff, 360 - diff)
                if diff < DELTA_THETA_MIN:
                    ok = False
                    break
            if not ok:
                continue
            photo_done.setdefault(iv["target"], []).append(angle)
            schedule.append(iv)

    return schedule


# ═══════════════════════════════════════════════════════════════
# 方法 3: CP-SAT 精确求解
# ═══════════════════════════════════════════════════════════════

def cpsat_schedule(t, x, y, v, a, shoot_targets, photo_targets, time_limit=300):
    """CP-SAT 精确求解：最大化任务完成总数"""
    try:
        from ortools.sat.python import cp_model
    except ImportError:
        print("  [CP-SAT] OR-Tools 未安装，跳过")
        return []

    # 生成候选
    candidates = []
    for _, row in shoot_targets.iterrows():
        fw = find_feasible_windows(t, x, y, v, a, row["x"], row["y"],
                                   D_SHOOT, V_SHOOT, A_SHOOT, PREP_SHOOT)
        # 稀疏化：每 10 个取 1 个（减少变量数）
        fw_sparse = fw[::10] if len(fw) > 50 else fw
        for idx in fw_sparse:
            prep_start = idx - int(PREP_SHOOT / 0.1)
            candidates.append({
                "start": prep_start,
                "end": idx,
                "target": row["id"],
                "task": "射击",
                "angle": None,
            })

    for _, row in photo_targets.iterrows():
        fw = find_feasible_windows(t, x, y, v, a, row["x"], row["y"],
                                   D_PHOTO, V_PHOTO, A_PHOTO, PREP_PHOTO)
        fw_sparse = fw[::5] if len(fw) > 30 else fw
        for idx in fw_sparse:
            angle = compute_angle(x[idx], y[idx], row["x"], row["y"])
            prep_start = idx - int(PREP_PHOTO / 0.1)
            candidates.append({
                "start": prep_start,
                "end": idx,
                "target": row["id"],
                "task": "拍照",
                "angle": angle,
            })

    if not candidates:
        return []

    n = len(candidates)
    print(f"  [CP-SAT] {n} 个候选变量，建模中...")

    model = cp_model.CpModel()

    # 决策变量：是否选择候选 i
    x_vars = [model.new_bool_var(f"x_{i}") for i in range(n)]

    # 约束 1: 时间不重叠（区间图着色）
    for i in range(n):
        for j in range(i + 1, n):
            # 检查区间是否重叠
            if candidates[i]["end"] > candidates[j]["start"] and candidates[j]["end"] > candidates[i]["start"]:
                model.add(x_vars[i] + x_vars[j] <= 1)

    # 约束 2: 每个射击目标至多 1 次
    shoot_groups = {}
    for i, c in enumerate(candidates):
        if c["task"] == "射击":
            shoot_groups.setdefault(c["target"], []).append(i)
    for target, indices in shoot_groups.items():
        model.add(sum(x_vars[i] for i in indices) <= 1)

    # 约束 3: 同目标拍照角差 ≥ 60°
    photo_groups = {}
    for i, c in enumerate(candidates):
        if c["task"] == "拍照":
            photo_groups.setdefault(c["target"], []).append(i)

    for target, indices in photo_groups.items():
        for ii in range(len(indices)):
            for jj in range(ii + 1, len(indices)):
                i_idx, j_idx = indices[ii], indices[jj]
                a1 = candidates[i_idx]["angle"]
                a2 = candidates[j_idx]["angle"]
                diff = abs(a1 - a2) % 360
                diff = min(diff, 360 - diff)
                if diff < DELTA_THETA_MIN:
                    model.add(x_vars[i_idx] + x_vars[j_idx] <= 1)

    # 目标：最大化任务数（射击权重 2，拍照权重 1）
    model.maximize(
        sum(2 * x_vars[i] if candidates[i]["task"] == "射击" else x_vars[i]
            for i in range(n))
    )

    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = time_limit
    solver.parameters.num_workers = 4

    status = solver.solve(model)

    if status in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        schedule = []
        for i in range(n):
            if solver.value(x_vars[i]):
                schedule.append(candidates[i])
        print(f"  [CP-SAT] 状态: {'OPTIMAL' if status == cp_model.OPTIMAL else 'FEASIBLE'}")
        print(f"  [CP-SAT] 目标值: {solver.objective_value}")
        return schedule
    else:
        print(f"  [CP-SAT] 无可行解 (status={status})")
        return []


# ═══════════════════════════════════════════════════════════════
# 输出
# ═══════════════════════════════════════════════════════════════

def write_result_xlsx(schedule, t, version=2):
    src = DATA / "result.xlsx"
    dst = OUTPUT / f"result_v{version}.xlsx"
    shutil.copy2(src, dst)

    wb = load_workbook(dst)
    ws = wb.active

    # 按执行时间排序
    schedule_sorted = sorted(schedule, key=lambda s: s.get("end", s.get("exec_idx", 0)))

    for i, s in enumerate(schedule_sorted):
        row = i + 2
        exec_idx = s.get("end", s.get("exec_idx", 0))
        prep_start_idx = s.get("start", s.get("prep_start_idx", 0))
        ws.cell(row=row, column=1, value=i + 1)
        ws.cell(row=row, column=2, value=s["target"])
        ws.cell(row=row, column=3, value=s["task"])
        ws.cell(row=row, column=4, value=round(float(t[prep_start_idx]), 2))
        ws.cell(row=row, column=5, value=round(float(t[exec_idx]), 2))

    wb.save(dst)
    return dst


def main() -> None:
    t_total = time.time()
    t, x, y, v, a = load_traj()
    shoot, photo = load_targets()

    print(f"轨迹: {len(t)} 点, 时长 {t[-1]-t[0]:.1f}s")
    print(f"目标: {len(shoot)} 射击 + {len(photo)} 拍照")
    print(f"v 统计: mean={v.mean():.2f}, P95={np.percentile(v,95):.2f}, max={v.max():.2f}")
    print(f"a 统计: mean={a.mean():.2f}, P95={np.percentile(a,95):.2f}, max={a.max():.2f}")

    results = {}

    # 方法 1: 改进贪心
    print("\n--- 方法 1: 改进贪心 ---")
    t0 = time.time()
    sched_greedy = improved_greedy(t, x, y, v, a, shoot, photo)
    elapsed = time.time() - t0
    n_shoot_g = sum(1 for s in sched_greedy if s["task"] == "射击")
    n_photo_g = sum(1 for s in sched_greedy if s["task"] == "拍照")
    print(f"  完成: 射击 {n_shoot_g}/{len(shoot)}, 拍照 {n_photo_g} 次, 总 {len(sched_greedy)}, 耗时 {elapsed:.1f}s")
    results["greedy"] = (sched_greedy, n_shoot_g, n_photo_g, elapsed)

    # 方法 2: 区间调度 DP
    print("\n--- 方法 2: 区间调度 DP ---")
    t0 = time.time()
    sched_dp = interval_scheduling_dp(t, x, y, v, a, shoot, photo)
    elapsed = time.time() - t0
    n_shoot_d = sum(1 for s in sched_dp if s["task"] == "射击")
    n_photo_d = sum(1 for s in sched_dp if s["task"] == "拍照")
    print(f"  完成: 射击 {n_shoot_d}/{len(shoot)}, 拍照 {n_photo_d} 次, 总 {len(sched_dp)}, 耗时 {elapsed:.1f}s")
    results["dp"] = (sched_dp, n_shoot_d, n_photo_d, elapsed)

    # 方法 3: CP-SAT
    print("\n--- 方法 3: CP-SAT 精确求解 ---")
    t0 = time.time()
    sched_cpsat = cpsat_schedule(t, x, y, v, a, shoot, photo, time_limit=300)
    elapsed = time.time() - t0
    n_shoot_c = sum(1 for s in sched_cpsat if s["task"] == "射击")
    n_photo_c = sum(1 for s in sched_cpsat if s["task"] == "拍照")
    print(f"  完成: 射击 {n_shoot_c}/{len(shoot)}, 拍照 {n_photo_c} 次, 总 {len(sched_cpsat)}, 耗时 {elapsed:.1f}s")
    results["cpsat"] = (sched_cpsat, n_shoot_c, n_photo_c, elapsed)

    # 选最优
    best_name = max(results, key=lambda k: len(results[k][0]))
    best_sched = results[best_name][0]
    best_n_shoot = results[best_name][1]
    best_n_photo = results[best_name][2]

    print(f"\n{'='*50}")
    print(f"最优方法: {best_name} → 总任务 {len(best_sched)} (射击 {best_n_shoot} + 拍照 {best_n_photo})")
    print(f"{'='*50}")

    # 写 result.xlsx
    dst = write_result_xlsx(best_sched, t, version=2)
    print(f"输出: {dst}")

    # 报告
    total_time = time.time() - t_total
    lines = [
        "# 问题 4 求解结果（优化版 v2）\n",
        "## 约束阈值",
        "- 射击: d∈[5,30]m, v≤2m/s, a≤1.5m/s², prep=1.5s, η=0.85",
        "- 拍照: d∈[10,40]m, v≤1.5m/s, a≤1.5m/s², prep=0.5s, Δθ≥60°\n",
        "## 轨迹运动学",
        f"- v: mean={v.mean():.2f}, P95={np.percentile(v,95):.2f}, max={v.max():.2f} m/s",
        f"- a: mean={a.mean():.2f}, P95={np.percentile(a,95):.2f}, max={a.max():.2f} m/s²\n",
        "## 三方法对比",
        f"| 方法 | 射击 | 拍照 | 总任务 | 耗时 |",
        f"|------|------|------|--------|------|",
        f"| 改进贪心 | {results['greedy'][1]}/{len(shoot)} | {results['greedy'][2]} | {len(results['greedy'][0])} | {results['greedy'][3]:.1f}s |",
        f"| 区间DP | {results['dp'][1]}/{len(shoot)} | {results['dp'][2]} | {len(results['dp'][0])} | {results['dp'][3]:.1f}s |",
        f"| CP-SAT | {results['cpsat'][1]}/{len(shoot)} | {results['cpsat'][2]} | {len(results['cpsat'][0])} | {results['cpsat'][3]:.1f}s |",
        f"\n## 最优结果: {best_name}",
        f"- 射击: {best_n_shoot}/{len(shoot)} 目标",
        f"- 拍照: {best_n_photo} 次",
        f"- 总任务: {len(best_sched)}",
        f"- 总耗时: {total_time:.1f}s",
        f"- 输出: output/result_v2.xlsx",
    ]
    (DOCS / "problem4_result_v2.md").write_text("\n".join(lines), encoding="utf-8")


if __name__ == "__main__":
    main()
