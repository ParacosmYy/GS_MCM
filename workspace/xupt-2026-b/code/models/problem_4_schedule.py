"""
问题 4：任务规划（射击 + 拍照）

约束（从 WMF 提取确认）：
  射击: d ∈ [5, 30] m, v ≤ 2 m/s, a ≤ 1.5 m/s², prep=1.5s
  拍照: d ∈ [10, 40] m, v ≤ 1.5 m/s, a ≤ 1.5 m/s², prep=0.5s, Δθ ≥ 60°

策略（三路）：
1. 贪心基线（2 min 内必出解）
2. CP-SAT（20 min wall-clock 上限）
3. 取得分最高的结果输出 result.xlsx

输入：output/traj_10hz_3.csv（3691 点 10Hz）
输出：output/result_v1.xlsx, docs/problem4_result.md
"""
from __future__ import annotations

import json
import time
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
DATA = ROOT / "data"
OUTPUT = ROOT / "output"
DOCS = ROOT / "docs"
OUTPUT.mkdir(exist_ok=True)

# ─── 约束常量 ───
D_SHOOT = (5, 30)
V_SHOOT = 2.0
A_SHOOT = 1.5
PREP_SHOOT = 1.5

D_PHOTO = (10, 40)
V_PHOTO = 1.5
A_PHOTO = 1.5
PREP_PHOTO = 0.5
DELTA_THETA_MIN = 60  # degrees


def load_traj():
    from scipy.signal import savgol_filter
    traj = pd.read_csv(OUTPUT / "traj_10hz_3.csv")
    t = traj["t"].to_numpy()
    # SG 滤波平滑位置（窗口 101 点 = 10s，3 阶多项式）
    # 真实数据噪声 σ≈3-4m，需较大窗口才能得到合理的 v/a
    window = min(101, len(t) - (1 if len(t) % 2 == 0 else 0))
    if window % 2 == 0:
        window -= 1
    x = savgol_filter(traj["x"].to_numpy(), window, 3)
    y = savgol_filter(traj["y"].to_numpy(), window, 3)
    dt = 0.1
    # SG 也可直接输出导数，但这里用平滑后差分更直观
    vx = np.gradient(x, dt)
    vy = np.gradient(y, dt)
    v = np.sqrt(vx**2 + vy**2)
    ax = np.gradient(vx, dt)
    ay = np.gradient(vy, dt)
    a = np.sqrt(ax**2 + ay**2)
    return t, x, y, v, a


def load_targets():
    shoot = pd.read_excel(DATA / "附件4.xlsx", sheet_name="射击目标")
    photo = pd.read_excel(DATA / "附件4.xlsx", sheet_name="拍照目标")
    shoot.columns = ["id", "x", "y"]
    photo.columns = ["id", "x", "y"]
    return shoot, photo


def find_feasible_windows(t, x, y, v, a, target_x, target_y, d_range, v_max, a_max, prep_time):
    """找出所有可行执行时刻（满足 prep 窗口内约束持续成立）"""
    n = len(t)
    dt = 0.1
    prep_samples = int(prep_time / dt)  # 1.5s → 15 samples, 0.5s → 5 samples

    dist = np.sqrt((x - target_x)**2 + (y - target_y)**2)
    d_ok = (dist >= d_range[0]) & (dist <= d_range[1])
    v_ok = v <= v_max
    a_ok = a <= a_max
    instant_ok = d_ok & v_ok & a_ok

    feasible = []
    for i in range(prep_samples, n):
        if np.all(instant_ok[i - prep_samples:i + 1]):
            feasible.append(i)
    return feasible


def greedy_schedule(t, x, y, v, a, shoot_targets, photo_targets):
    """贪心策略：按时间顺序扫描，每遇到可行窗口就执行"""
    schedule = []  # (time_idx, target_id, task_type, prep_start_t, exec_t)
    occupied_until = -1  # 时间轴占用标记

    # 预计算所有目标的可行窗
    candidates = []

    for _, row in shoot_targets.iterrows():
        fw = find_feasible_windows(t, x, y, v, a, row["x"], row["y"],
                                   D_SHOOT, V_SHOOT, A_SHOOT, PREP_SHOOT)
        for idx in fw:
            candidates.append((idx, row["id"], "射击", PREP_SHOOT))

    for _, row in photo_targets.iterrows():
        fw = find_feasible_windows(t, x, y, v, a, row["x"], row["y"],
                                   D_PHOTO, V_PHOTO, A_PHOTO, PREP_PHOTO)
        for idx in fw:
            angle = np.degrees(np.arctan2(y[idx] - row["y"], x[idx] - row["x"]))
            candidates.append((idx, row["id"], "拍照", PREP_PHOTO, angle))

    # 按执行时刻排序
    candidates.sort(key=lambda c: c[0])

    # 贪心：射击每目标至多 1 次；拍照同目标多次需角差 ≥ 60°
    shot_done = set()
    photo_done = {}  # target_id → list of angles

    for c in candidates:
        idx = c[0]
        target_id = c[1]
        task = c[2]
        prep = c[3]

        prep_samples = int(prep / 0.1)
        prep_start_idx = idx - prep_samples
        # 检查时间窗不冲突
        if prep_start_idx <= occupied_until:
            continue

        if task == "射击":
            if target_id in shot_done:
                continue
            shot_done.add(target_id)
            schedule.append({
                "target": target_id,
                "task": "射击",
                "prep_start_t": float(t[prep_start_idx]),
                "exec_t": float(t[idx]),
            })
            occupied_until = idx

        elif task == "拍照":
            angle = c[4]
            prev_angles = photo_done.get(target_id, [])
            # 检查角差（处理 ±180 度环绕）
            ok = True
            for pa in prev_angles:
                diff = abs(angle - pa) % 360
                diff = min(diff, 360 - diff)
                if diff < DELTA_THETA_MIN:
                    ok = False
                    break
            if not ok:
                continue
            photo_done.setdefault(target_id, []).append(angle)
            schedule.append({
                "target": target_id,
                "task": "拍照",
                "prep_start_t": float(t[prep_start_idx]),
                "exec_t": float(t[idx]),
            })
            occupied_until = idx

    return schedule


def write_result_xlsx(schedule, version=1):
    """写 result.xlsx 保留红字"""
    # 从模板复制
    import shutil
    src = DATA / "result.xlsx"
    dst = OUTPUT / f"result_v{version}.xlsx"
    shutil.copy2(src, dst)

    wb = load_workbook(dst)
    ws = wb.active

    # 按时间排序
    schedule.sort(key=lambda s: s["exec_t"])

    for i, s in enumerate(schedule):
        row = i + 2  # A1 是表头（红字保留）
        ws.cell(row=row, column=1, value=i + 1)
        ws.cell(row=row, column=2, value=s["target"])
        ws.cell(row=row, column=3, value=s["task"])
        ws.cell(row=row, column=4, value=round(s["prep_start_t"], 2))
        ws.cell(row=row, column=5, value=round(s["exec_t"], 2))

    wb.save(dst)
    return dst


def main() -> None:
    t0 = time.time()
    t, x, y, v, a = load_traj()
    shoot, photo = load_targets()

    print(f"轨迹: {len(t)} 点, 时长 {t[-1]-t[0]:.1f}s")
    print(f"目标: {len(shoot)} 射击 + {len(photo)} 拍照")
    print(f"v 统计: mean={v.mean():.2f}, P95={np.percentile(v,95):.2f}, max={v.max():.2f}")
    print(f"a 统计: mean={a.mean():.2f}, P95={np.percentile(a,95):.2f}, max={a.max():.2f}")

    # 贪心
    print("\n--- 贪心求解 ---")
    schedule = greedy_schedule(t, x, y, v, a, shoot, photo)
    elapsed = time.time() - t0
    n_shoot = sum(1 for s in schedule if s["task"] == "射击")
    n_photo = sum(1 for s in schedule if s["task"] == "拍照")
    print(f"贪心完成: {elapsed:.1f}s, 射击 {n_shoot}/{len(shoot)}, 拍照 {n_photo} 次")

    # 写 result
    dst = write_result_xlsx(schedule, version=1)
    print(f"输出: {dst}")

    # 报告
    lines = [
        "# 问题 4 求解结果\n",
        "## 约束阈值（确认）",
        f"- 射击: d∈[5,30]m, v≤2m/s, a≤1.5m/s², prep=1.5s, η=0.85",
        f"- 拍照: d∈[10,40]m, v≤1.5m/s, a≤1.5m/s², prep=0.5s, Δθ≥60°\n",
        "## 轨迹运动学",
        f"- v: mean={v.mean():.2f}, P95={np.percentile(v,95):.2f}, max={v.max():.2f} m/s",
        f"- a: mean={a.mean():.2f}, P95={np.percentile(a,95):.2f}, max={a.max():.2f} m/s²\n",
        "## 贪心基线结果",
        f"- 射击完成: {n_shoot}/{len(shoot)} 目标",
        f"- 拍照完成: {n_photo} 次（跨 {len(set(s['target'] for s in schedule if s['task']=='拍照'))} 个目标）",
        f"- 总任务数: {len(schedule)}",
        f"- 求解时间: {elapsed:.1f}s",
        f"- 输出: output/result_v1.xlsx\n",
        "## 任务明细（前 10 条）",
    ]
    for s in schedule[:10]:
        lines.append(f"  {s['target']}  {s['task']}  prep@{s['prep_start_t']:.2f}  exec@{s['exec_t']:.2f}")
    if len(schedule) > 10:
        lines.append(f"  ... 共 {len(schedule)} 条")

    (DOCS / "problem4_result.md").write_text("\n".join(lines), encoding="utf-8")
    print(f"\n总任务: {len(schedule)}  (射击 {n_shoot} + 拍照 {n_photo})")


if __name__ == "__main__":
    main()
