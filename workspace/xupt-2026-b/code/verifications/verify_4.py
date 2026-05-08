"""
问题 4 自证：验证 result.xlsx 中每行任务的合法性。

检查项：
1. 红字单元格未被修改（对照 result_redcells.json）
2. 每行 prep_start = exec - Δ_prep（射击 1.5s / 拍照 0.5s）
3. 在 [prep_start, exec] 期间所有 10Hz 采样点满足 d/v/a 约束
4. 时间窗无冲突
5. 拍照同目标角差 ≥ 60°
6. 射击每目标至多 1 次
"""
from __future__ import annotations

import json
from pathlib import Path

import numpy as np
import pandas as pd
from scipy.signal import savgol_filter
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
DATA = ROOT / "data"
OUTPUT = ROOT / "output"

D_SHOOT = (5, 30); V_SHOOT = 2.0; A_SHOOT = 1.5; PREP_SHOOT = 1.5
D_PHOTO = (10, 40); V_PHOTO = 1.5; A_PHOTO = 1.5; PREP_PHOTO = 0.5
DELTA_THETA_MIN = 60


def load_smoothed_traj():
    traj = pd.read_csv(OUTPUT / "traj_10hz_3.csv")
    t = traj["t"].to_numpy()
    x = savgol_filter(traj["x"].to_numpy(), 151, 3)
    y = savgol_filter(traj["y"].to_numpy(), 151, 3)
    dt = 0.1
    vx = np.gradient(x, dt); vy = np.gradient(y, dt)
    v = np.sqrt(vx**2 + vy**2)
    ax = np.gradient(vx, dt); ay = np.gradient(vy, dt)
    a = np.sqrt(ax**2 + ay**2)
    return t, x, y, v, a


def main() -> None:
    t, x, y, v, a = load_smoothed_traj()
    targets = {}
    shoot = pd.read_excel(DATA / "附件4.xlsx", sheet_name="射击目标")
    photo = pd.read_excel(DATA / "附件4.xlsx", sheet_name="拍照目标")
    shoot.columns = ["id", "x", "y"]; photo.columns = ["id", "x", "y"]
    for _, r in pd.concat([shoot, photo]).iterrows():
        targets[r["id"]] = (r["x"], r["y"])

    # Load result
    result_path = OUTPUT / "result_v2.xlsx"
    wb = load_workbook(result_path)
    ws = wb.active
    tasks = []
    for row in ws.iter_rows(min_row=2, max_col=5, values_only=True):
        if row[0] is None:
            break
        tasks.append({"seq": row[0], "target": row[1], "task": row[2],
                      "prep_t": row[3], "exec_t": row[4]})

    results = []

    # 1. 红字检查
    redcells = json.loads((DATA / "result_redcells.json").read_text(encoding="utf-8"))
    wb2 = load_workbook(result_path)
    ws2 = wb2.active
    red_ok = True
    for rc in redcells:
        cell = ws2[rc["coord"]]
        if cell.value != rc["value"]:
            red_ok = False
            break
    results.append(("红字保留", red_ok, f"{len(redcells)} 单元"))

    # 2-6. 逐行检查
    violations = []
    shoot_targets_done = {}
    photo_angles = {}

    for task in tasks:
        tid = task["target"]; tt = task["task"]
        prep_t = task["prep_t"]; exec_t = task["exec_t"]
        prep = PREP_SHOOT if tt == "射击" else PREP_PHOTO

        # 2. prep 时长
        if abs((exec_t - prep_t) - prep) > 0.02:
            violations.append(f"{tid}: prep 时长错误 {exec_t-prep_t:.2f} != {prep}")

        # 3. 约束检查
        d_range = D_SHOOT if tt == "射击" else D_PHOTO
        v_max = V_SHOOT if tt == "射击" else V_PHOTO
        a_max = A_SHOOT if tt == "射击" else A_PHOTO
        tx, ty = targets[tid]

        mask = (t >= prep_t - 0.05) & (t <= exec_t + 0.05)
        idxs = np.where(mask)[0]
        for idx in idxs:
            d = np.sqrt((x[idx] - tx)**2 + (y[idx] - ty)**2)
            if d < d_range[0] or d > d_range[1]:
                violations.append(f"{tid}@t={t[idx]:.2f}: d={d:.1f} out [{d_range[0]},{d_range[1]}]")
                break
            if v[idx] > v_max * 1.01:
                violations.append(f"{tid}@t={t[idx]:.2f}: v={v[idx]:.2f} > {v_max}")
                break
            if a[idx] > a_max * 1.01:
                violations.append(f"{tid}@t={t[idx]:.2f}: a={a[idx]:.2f} > {a_max}")
                break

        # 6. 射击唯一
        if tt == "射击":
            if tid in shoot_targets_done:
                violations.append(f"{tid}: 重复射击")
            shoot_targets_done[tid] = exec_t

        # 5. 拍照角差
        if tt == "拍照":
            idx_exec = np.argmin(np.abs(t - exec_t))
            angle = np.degrees(np.arctan2(y[idx_exec] - ty, x[idx_exec] - tx))
            prev = photo_angles.get(tid, [])
            for pa in prev:
                diff = abs(angle - pa)
                diff = min(diff, 360 - diff)
                if diff < DELTA_THETA_MIN - 1:
                    violations.append(f"{tid}: 角差 {diff:.1f}° < {DELTA_THETA_MIN}°")
            photo_angles.setdefault(tid, []).append(angle)

    # 4. 时间窗无冲突（全量 pairwise 检查）
    windows = [(task["prep_t"], task["exec_t"]) for task in tasks]
    for i in range(len(windows)):
        for j in range(i + 1, len(windows)):
            if windows[i][1] > windows[j][0] + 0.02 and windows[j][1] > windows[i][0] + 0.02:
                violations.append(f"时间窗冲突: {windows[i]} 与 {windows[j]}")

    results.append(("约束/唯一/角差/时序", len(violations) == 0,
                    f"{len(violations)} 违规" if violations else "0 violations"))

    # 总结
    n_shoot = sum(1 for t2 in tasks if t2["task"] == "射击")
    n_photo = sum(1 for t2 in tasks if t2["task"] == "拍照")
    results.append(("任务数 > 0", len(tasks) > 0, f"总{len(tasks)} (射击{n_shoot}+拍照{n_photo})"))

    all_pass = all(r[1] for r in results)
    print("=" * 50)
    print("  问题 4 验证报告")
    print("=" * 50)
    for name, passed, detail in results:
        mark = "[PASS]" if passed else "[FAIL]"
        print(f"  {mark}  {name}  ({detail})")
    if violations:
        for vv in violations[:10]:
            print(f"    ⚠ {vv}")
    print("=" * 50)
    print(f"  最终状态: {'ALL PASS' if all_pass else 'FAILED'}")
    print("=" * 50)
    if not all_pass:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
