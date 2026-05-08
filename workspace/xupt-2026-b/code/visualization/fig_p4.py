#!/usr/bin/env python3
"""
Generate 4 high-quality figures for Problem 4 analysis
- Mission map with trajectories and targets
- Gantt chart for task timing
- Velocity/acceleration profiles with feasibility windows
- Score pie chart
"""

import matplotlib
matplotlib.use('Agg')  # Non-interactive backend

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from scipy.signal import savgol_filter
from pathlib import Path
import openpyxl

# Configuration
BASE_DIR = Path(__file__).resolve().parents[2]
TRAJ_FILE = BASE_DIR / "output/traj_10hz_3.csv"
TARGET_FILE = BASE_DIR / "data/附件4.xlsx"
RESULT_FILE = BASE_DIR / "output/result_v2.xlsx"
FIG_DIR = BASE_DIR / "figures"

# Constraint constants
D_SHOOT = (5, 30)
V_SHOOT = 2.0
A_SHOOT = 1.5
PREP_SHOOT = 1.5

D_PHOTO = (10, 40)
V_PHOTO = 1.5
A_PHOTO = 1.5
PREP_PHOTO = 0.5

# Styling
DPI = 200
plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'DejaVu Sans']
plt.rcParams['axes.unicode_minus'] = False
plt.rcParams['font.size'] = 10
plt.rcParams['axes.labelsize'] = 11
plt.rcParams['axes.titlesize'] = 12
plt.rcParams['xtick.labelsize'] = 9
plt.rcParams['ytick.labelsize'] = 9
plt.rcParams['legend.fontsize'] = 9
plt.rcParams['figure.titlesize'] = 13


def load_trajectory(filepath, window=151):
    """Load trajectory and apply Savitzky-Golay smoothing"""
    df = pd.read_csv(filepath)
    t = df['t'].values
    x = df['x'].values
    y = df['y'].values

    # Apply SG filter
    if len(t) >= window:
        x_smooth = savgol_filter(x, window, 3)
        y_smooth = savgol_filter(y, window, 3)
    else:
        x_smooth = x
        y_smooth = y

    return t, x_smooth, y_smooth


def load_targets(filepath):
    """Load shooting and photo targets from Excel"""
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # Shooting targets
    ws_shoot = wb["射击目标"]
    shoot_targets = []
    for row in ws_shoot.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break
        shoot_targets.append({
            'id': row[0],
            'x': float(row[1]),
            'y': float(row[2])
        })

    # Photo targets
    ws_photo = wb["拍照目标"]
    photo_targets = []
    for row in ws_photo.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break
        photo_targets.append({
            'id': row[0],
            'x': float(row[1]),
            'y': float(row[2])
        })

    return shoot_targets, photo_targets


def load_results(filepath):
    """Load task execution results from Excel"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    results = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break
        results.append({
            'seq': row[0],
            'target_id': row[1],
            'task': row[2],
            'prep_start': float(row[3]) if row[3] is not None else None,
            'exec_time': float(row[4]) if row[4] is not None else None
        })

    return results


def compute_velocity_acceleration(t, x, y):
    """Compute velocity and acceleration from position"""
    dt = np.diff(t)
    dx = np.diff(x)
    dy = np.diff(y)

    vx = dx / dt
    vy = dy / dt
    v = np.sqrt(vx**2 + vy**2)

    # Pad to match length
    v = np.concatenate([[v[0]], v])

    # Acceleration
    dv = np.diff(v)
    a = dv / dt
    a = np.concatenate([[a[0]], a])

    return v, np.abs(a)


def find_robot_position(t_array, x_array, y_array, target_time):
    """Find robot position at target time"""
    idx = np.argmin(np.abs(t_array - target_time))
    return x_array[idx], y_array[idx]


# ============================================================================
# Figure 1: Mission Map
# ============================================================================
def plot_mission_map(t, x, y, shoot_targets, photo_targets, results):
    """Plot trajectory and targets with execution indicators"""
    fig, ax = plt.subplots(figsize=(10, 8), dpi=DPI)

    # Plot trajectory colored by velocity
    from matplotlib.collections import LineCollection
    v_local = np.sqrt(np.gradient(x, 0.1)**2 + np.gradient(y, 0.1)**2)
    points = np.array([x, y]).T.reshape(-1, 1, 2)
    segments = np.concatenate([points[:-1], points[1:]], axis=1)
    lc = LineCollection(segments, cmap='coolwarm', linewidth=2.5, alpha=0.7)
    lc.set_array(v_local[:-1])
    lc.set_clim(0, 5)
    ax.add_collection(lc)
    cbar = plt.colorbar(lc, ax=ax, shrink=0.6, pad=0.02)
    cbar.set_label('速度 (m/s)', fontsize=10)

    # Plot all targets
    shoot_ids = [t['id'] for t in shoot_targets]
    photo_ids = [t['id'] for t in photo_targets]

    for target in shoot_targets:
        ax.plot(target['x'], target['y'], '^', color='#C0392B', markersize=11,
                markeredgecolor='darkred', markeredgewidth=1, alpha=0.8)

    for target in photo_targets:
        ax.plot(target['x'], target['y'], 'o', color='#2980B9', markersize=11,
                markeredgecolor='darkblue', markeredgewidth=1, alpha=0.8)

    # Mark successful executions
    executed_shoot = set()
    executed_photo = set()

    for res in results:
        if res['exec_time'] is None:
            continue

        target_id = res['target_id']
        task_type = res['task']

        # Find target coordinates
        target_coord = None
        if '射击' in task_type:
            for t_obj in shoot_targets:
                if t_obj['id'] == target_id:
                    target_coord = (t_obj['x'], t_obj['y'])
                    break
            executed_shoot.add(target_id)
        elif '拍照' in task_type:
            for t_obj in photo_targets:
                if t_obj['id'] == target_id:
                    target_coord = (t_obj['x'], t_obj['y'])
                    break
            executed_photo.add(target_id)

        if target_coord:
            # Find robot position at execution time
            robot_x, robot_y = find_robot_position(t, x, y, res['exec_time'])

            # Draw dashed line from robot to target
            ax.plot([robot_x, target_coord[0]], [robot_y, target_coord[1]],
                   '--', color='green', linewidth=1, alpha=0.5)

            # Overlay success marker
            if '射击' in task_type:
                ax.plot(target_coord[0], target_coord[1], '^',
                       color='green', markersize=12, markeredgecolor='darkgreen',
                       markeredgewidth=1.5)
            else:
                ax.plot(target_coord[0], target_coord[1], 'o',
                       color='green', markersize=12, fillstyle='none',
                       markeredgecolor='green', markeredgewidth=2)

    # Custom legend
    from matplotlib.lines import Line2D
    legend_elements = [
        Line2D([0], [0], color='gray', linewidth=2, alpha=0.4, label='机器人轨迹'),
        Line2D([0], [0], marker='^', color='w', markerfacecolor='red',
               markersize=8, label='射击目标'),
        Line2D([0], [0], marker='o', color='w', markerfacecolor='blue',
               markersize=8, label='拍照目标'),
        Line2D([0], [0], marker='^', color='w', markerfacecolor='green',
               markersize=10, label='射击完成'),
        Line2D([0], [0], marker='o', color='w', markerfacecolor='none',
               markeredgecolor='green', markersize=10, markeredgewidth=2,
               label='拍照完成'),
    ]
    ax.legend(handles=legend_elements, loc='best')

    ax.set_xlabel('X (m)', fontsize=11)
    ax.set_ylabel('Y (m)', fontsize=11)
    ax.set_title('任务地图：轨迹与目标执行情况', fontsize=13, weight='bold')
    ax.set_aspect('equal')
    ax.grid(True, alpha=0.2, linestyle='--')

    plt.tight_layout()
    plt.savefig(FIG_DIR / "fig_p4_mission_map.png", dpi=DPI)
    plt.close()

    print(f"[OK] Saved: fig_p4_mission_map.png")


# ============================================================================
# Figure 2: Gantt Chart
# ============================================================================
def plot_gantt_chart(results):
    """Plot Gantt chart of task execution timeline — enhanced version"""
    valid = [r for r in results if r['prep_start'] is not None and r['exec_time'] is not None]
    n_tasks = len(valid)
    fig, ax = plt.subplots(figsize=(14, max(8, n_tasks * 0.35)), dpi=DPI)

    colors = {'shoot': '#C0392B', 'photo': '#2980B9'}

    for y_pos, res in enumerate(valid):
        is_shoot = '射' in str(res['task']) or str(res['target_id']).startswith('S')
        color = colors['shoot'] if is_shoot else colors['photo']
        task_type = 'Shoot' if is_shoot else 'Photo'

        prep_dur = res['exec_time'] - res['prep_start']
        ax.barh(y_pos, prep_dur, left=res['prep_start'], height=0.75,
                color=color, alpha=0.85, edgecolor='white', linewidth=0.8)

        ax.text(res['prep_start'] + prep_dur / 2, y_pos,
                f"{res['target_id']}", ha='center', va='center',
                fontsize=9, color='white', weight='bold')

    ax.set_yticks(range(n_tasks))
    ax.set_yticklabels([f"{r['target_id']} ({('S' if '射' in str(r['task']) or str(r['target_id']).startswith('S') else 'P')})"
                        for r in valid], fontsize=9)
    ax.set_xlabel('时间 (s)', fontsize=11)
    ax.set_title(f'任务执行时间线 ({n_tasks} Tasks)', fontsize=13, weight='bold')
    ax.grid(True, axis='x', alpha=0.25, linestyle='--')
    ax.grid(True, axis='y', alpha=0.1)
    ax.invert_yaxis()
    ax.set_xlim(left=min(r['prep_start'] for r in valid) - 5)

    from matplotlib.patches import Patch
    legend_elements = [
        Patch(facecolor=colors['shoot'], alpha=0.85, label=f'射击 ({sum(1 for r in valid if "射" in str(r["task"]) or str(r["target_id"]).startswith("S"))})'),
        Patch(facecolor=colors['photo'], alpha=0.85, label=f'拍照 ({sum(1 for r in valid if not ("射" in str(r["task"]) or str(r["target_id"]).startswith("S")))})'),
    ]
    ax.legend(handles=legend_elements, loc='lower right', fontsize=10, framealpha=0.9)

    plt.tight_layout()
    plt.savefig(FIG_DIR / "fig_p4_gantt.png", dpi=DPI, bbox_inches='tight')
    plt.close()
    print(f"[OK] Saved: fig_p4_gantt.png ({n_tasks} tasks)")


# ============================================================================
# Figure 3: Velocity/Acceleration Profile
# ============================================================================
def plot_velocity_profile(t, v, a):
    """Plot velocity, acceleration, and feasibility windows"""
    fig, axes = plt.subplots(3, 1, figsize=(12, 9), dpi=DPI, sharex=True)

    # Velocity plot
    ax = axes[0]
    ax.plot(t, v, '-', color='darkblue', linewidth=1, label='速度')
    ax.axhline(V_SHOOT, color='red', linestyle='--', linewidth=1,
              label=f'射击阈值 (v={V_SHOOT} m/s)')
    ax.axhline(V_PHOTO, color='blue', linestyle='--', linewidth=1,
              label=f'拍照阈值 (v={V_PHOTO} m/s)')
    ax.set_ylabel('速度 (m/s)')
    ax.set_title('速度曲线')
    ax.legend(loc='upper right')
    ax.grid(True, alpha=0.3)

    # Acceleration plot
    ax = axes[1]
    ax.plot(t, a, '-', color='darkred', linewidth=1, label='加速度')
    ax.axhline(A_SHOOT, color='orange', linestyle='--', linewidth=1,
              label=f'阈值 (a={A_SHOOT} m/s$^2$)')
    ax.set_ylabel(r'加速度 (m/s$^2$)')
    ax.set_title('加速度曲线')
    ax.legend(loc='upper right')
    ax.grid(True, alpha=0.3)

    # Feasibility windows
    ax = axes[2]

    # Shooting feasible: v <= V_SHOOT and a <= A_SHOOT
    shoot_feasible = (v <= V_SHOOT) & (a <= A_SHOOT)
    photo_feasible = (v <= V_PHOTO) & (a <= A_PHOTO)

    # Plot as vertical spans
    for i in range(len(t)):
        if shoot_feasible[i]:
            ax.axvline(t[i], color='red', alpha=0.1, linewidth=0.5)
        if photo_feasible[i]:
            ax.axvline(t[i], color='blue', alpha=0.1, linewidth=0.5)

    ax.set_ylabel('可行性')
    ax.set_xlabel('时间 (s)')
    ax.set_title('任务可行窗口')
    ax.set_ylim(-0.5, 1.5)
    ax.set_yticks([0, 1])
    ax.set_yticklabels(['', ''])

    # Custom legend
    from matplotlib.patches import Patch
    legend_elements = [
        Patch(facecolor='red', alpha=0.3, label='射击可行'),
        Patch(facecolor='blue', alpha=0.3, label='拍照可行')
    ]
    ax.legend(handles=legend_elements, loc='upper right')
    ax.grid(True, alpha=0.3)

    plt.tight_layout()
    plt.savefig(FIG_DIR / "fig_p4_velocity_profile.png", dpi=DPI)
    plt.close()

    print(f"[OK] Saved: fig_p4_velocity_profile.png")


# ============================================================================
# Figure 4: Score Pie Chart
# ============================================================================
def plot_score_pie(results, shoot_targets, photo_targets):
    """Plot score breakdown as pie/donut chart"""
    # Count completions
    shoot_completed = set()
    photo_completed = set()

    for res in results:
        if res['exec_time'] is not None:
            if '射击' in res['task']:
                shoot_completed.add(res['target_id'])
            elif '拍照' in res['task']:
                photo_completed.add(res['target_id'])

    n_shoot_total = len(shoot_targets)
    n_photo_total = len(photo_targets)
    n_shoot_done = len(shoot_completed)
    n_shoot_undone = n_shoot_total - n_shoot_done
    n_photo_done = len(photo_completed)
    n_photo_undone = n_photo_total - n_photo_done

    # Count photo tasks (can be multiple per target)
    n_photo_tasks = sum(1 for r in results if '拍照' in r['task'] and r['exec_time'] is not None)

    fig, ax = plt.subplots(figsize=(10, 8), dpi=DPI)

    # Data for pie chart
    labels = [
        f'射击完成\n({n_shoot_done}个目标)',
        f'射击未完成\n({n_shoot_undone}个目标)',
        f'拍照完成\n({n_photo_tasks}项任务，{n_photo_done}个目标)',
        f'拍照未完成\n({n_photo_undone}个目标)'
    ]

    sizes = [n_shoot_done, n_shoot_undone, n_photo_tasks, n_photo_undone]
    colors = ['#2ecc71', '#95a5a6', '#3498db', '#ecf0f1']
    explode = (0.05, 0, 0.05, 0)

    # Create donut chart
    wedges, texts, autotexts = ax.pie(sizes, labels=labels, colors=colors,
                                       autopct='%1.1f%%', startangle=90,
                                       explode=explode, textprops={'fontsize': 10})

    # Make it a donut
    centre_circle = plt.Circle((0, 0), 0.70, fc='white')
    ax.add_artist(centre_circle)

    # Center text
    total_tasks = sum(1 for r in results if r['exec_time'] is not None)
    total_possible = n_shoot_total + n_photo_total
    ax.text(0, 0, f'{total_tasks}\n项任务\n已完成',
           ha='center', va='center', fontsize=14, weight='bold')

    ax.set_title('任务完成情况', fontsize=14, weight='bold', pad=20)

    plt.tight_layout()
    plt.savefig(FIG_DIR / "fig_p4_score_pie.png", dpi=DPI)
    plt.close()

    print(f"[OK] Saved: fig_p4_score_pie.png")


# ============================================================================
# Main
# ============================================================================
def main():
    print("=" * 70)
    print("Generating Problem 4 Visualizations")
    print("=" * 70)

    # Load data
    print("\n[1/5] Loading trajectory data...")
    t, x, y = load_trajectory(TRAJ_FILE, window=151)
    print(f"      Loaded {len(t)} trajectory points")

    print("\n[2/5] Loading target data...")
    shoot_targets, photo_targets = load_targets(TARGET_FILE)
    print(f"      Shooting targets: {len(shoot_targets)}")
    print(f"      Photo targets: {len(photo_targets)}")

    print("\n[3/5] Loading execution results...")
    results = load_results(RESULT_FILE)
    print(f"      Task records: {len(results)}")

    print("\n[4/5] Computing velocity and acceleration...")
    v, a = compute_velocity_acceleration(t, x, y)
    print(f"      Computed for {len(v)} points")

    print("\n[5/5] Generating figures...")
    print()

    # Generate all figures
    plot_mission_map(t, x, y, shoot_targets, photo_targets, results)
    plot_gantt_chart(results)
    plot_velocity_profile(t, v, a)
    plot_score_pie(results, shoot_targets, photo_targets)

    print("\n" + "=" * 70)
    print("All figures generated successfully!")
    print(f"Output directory: {FIG_DIR}")
    print("=" * 70)


if __name__ == "__main__":
    main()
